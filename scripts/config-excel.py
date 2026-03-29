#!/usr/bin/env python3

from __future__ import annotations

import argparse
import copy
import json
import sys
import tempfile
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:
    print(
        "ERROR openpyxl is required for Excel import/export. Install it with: pip install openpyxl",
        file=sys.stderr,
    )
    raise SystemExit(1) from exc


OVERVIEW_SHEET = "Overview"
RULES_SHEET = "Rules"
FORMATTING_SHEET = "Formatting"
CATEGORY_SHEET = "Category Standards"
EDITOR_NOTES_SHEET = "Editor Notes"
EXTRA_JSON_SHEET = "Extra JSON"
REFERENCE_SHEET = "Reference"
INSTRUCTIONS_SHEET = "Instructions"

PRESENCE_VALUES = ["required", "recommended", "discouraged", "forbidden"]
GEOMETRY_VALUES = ["point", "polygon"]
BOOLEAN_VALUES = ["true", "false"]
SEVERITY_VALUES = ["info", "warning", "error"]

OVERVIEW_FIELDS = [
    ("id", "Config ID"),
    ("type", "Config Type"),
    ("version", "Version"),
    ("extends", "Extends"),
    ("meta.name", "Meta Name"),
    ("meta.description", "Meta Description"),
    ("scope.country", "Scope Country"),
    ("scope.countries", "Scope Countries"),
    ("defaults.locale", "Default Locale"),
]

CATEGORY_COLUMNS = [
    "categoryId",
    "categoryGroup",
    "categoryType",
    "isDefined",
    "geometryRequired",
    "geometryRecommended",
    "geometryAllowed",
    "lockLevel",
    "cityInVenueName",
    "phone",
    "url",
    "openingHours",
    "navigationPoints",
    "externalProviderIds",
    "addressCity",
    "addressStreet",
    "addressHouseNumber",
    "servicesRequired",
    "servicesRecommended",
    "servicesDiscouraged",
    "servicesForbidden",
    "extraJson",
]


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(data, handle, indent=2, ensure_ascii=False)
        handle.write("\n")


def read_cell(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def parse_boolean(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "ja", "1"}:
        return True
    if normalized in {"false", "no", "nee", "0"}:
        return False
    raise ValueError(f'Invalid boolean value "{value}"')


def stringify_boolean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    return read_cell(value)


def parse_integer(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return int(str(value).strip())


def split_lines(value: Any) -> list[str]:
    text = read_cell(value)
    if text is None:
        return []
    parts = [part.strip() for part in text.replace("\r\n", "\n").split("\n")]
    return [part for part in parts if part]


def join_lines(values: list[str] | None) -> str | None:
    if not values:
        return None
    return "\n".join(values)


def parse_json_cell(value: Any) -> Any:
    text = read_cell(value)
    if text is None:
        return None
    return json.loads(text)


def encode_value(value: Any) -> tuple[str, str | None]:
    if value is None:
        return ("empty", None)
    if isinstance(value, bool):
        return ("boolean", stringify_boolean(value))
    if isinstance(value, int) and not isinstance(value, bool):
        return ("integer", str(value))
    if isinstance(value, float):
        return ("number", str(value))
    if isinstance(value, str):
        return ("string", value)
    if isinstance(value, list) and all(isinstance(item, str) for item in value):
        return ("lines", join_lines(value))
    return ("json", json.dumps(value, ensure_ascii=False))


def decode_value(encoding: str | None, value: Any) -> Any:
    mode = (encoding or "string").strip().lower()
    if mode == "empty":
        return None
    if mode == "string":
        return read_cell(value)
    if mode == "boolean":
        return parse_boolean(value)
    if mode == "integer":
        return parse_integer(value)
    if mode == "number":
        text = read_cell(value)
        return float(text) if text is not None else None
    if mode == "lines":
        return split_lines(value)
    if mode == "json":
        return parse_json_cell(value)
    raise ValueError(f'Unsupported encoding "{encoding}"')


def get_nested(data: dict[str, Any], path: str) -> Any:
    current: Any = data
    for segment in path.split("."):
        if not isinstance(current, dict) or segment not in current:
            return None
        current = current[segment]
    return current


def set_nested(data: dict[str, Any], path: str, value: Any) -> None:
    current = data
    segments = path.split(".")
    for segment in segments[:-1]:
        next_value = current.get(segment)
        if not isinstance(next_value, dict):
            next_value = {}
            current[segment] = next_value
        current = next_value
    current[segments[-1]] = value


def deep_merge(target: dict[str, Any], source: dict[str, Any]) -> dict[str, Any]:
    for key, value in source.items():
        if isinstance(value, dict) and isinstance(target.get(key), dict):
            deep_merge(target[key], value)
            continue
        target[key] = value
    return target


def delete_nested(data: dict[str, Any], path: str) -> None:
    current = data
    parents: list[tuple[dict[str, Any], str]] = []
    for segment in path.split("."):
        if not isinstance(current, dict) or segment not in current:
            return
        parents.append((current, segment))
        current = current[segment]
    parent, key = parents.pop()
    parent.pop(key, None)
    while parents:
        parent, key = parents.pop()
        child = parent.get(key)
        if isinstance(child, dict) and not child:
            parent.pop(key, None)
            continue
        break


def prune_empty(value: Any) -> Any:
    if isinstance(value, dict):
        pruned: dict[str, Any] = {}
        for key, child in value.items():
            result = prune_empty(child)
            if result is None:
                continue
            if isinstance(result, dict) and not result:
                continue
            if isinstance(result, list) and not result:
                continue
            pruned[key] = result
        return pruned
    if isinstance(value, list):
        pruned_list = [prune_empty(child) for child in value]
        return [child for child in pruned_list if child is not None]
    return value


def remove_managed_paths(config: dict[str, Any]) -> dict[str, Any]:
    remainder = copy.deepcopy(config)
    for path, _label in OVERVIEW_FIELDS:
        delete_nested(remainder, path)
    for path in ["rules", "formatting", "categoryStandards"]:
        delete_nested(remainder, path)
    return prune_empty(remainder) or {}


def load_sdk_values(root: Path) -> dict[str, Any]:
    return load_json(root / "reference" / "sdk-values.json")


def build_category_parents(sdk_values: dict[str, Any]) -> dict[str, str | None]:
    parents: dict[str, str | None] = {category: None for category in sdk_values["categoryIds"]}
    for main_category in sdk_values["mainCategories"]:
        parents[main_category] = None
    for main_category, sub_categories in sdk_values["subCategoriesByMainCategory"].items():
        for sub_category in sub_categories:
            parents[sub_category] = main_category
    parents["RESIDENTIAL"] = None
    return parents


def create_header_row(worksheet, headers: list[str]) -> None:
    worksheet.append(headers)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def auto_fit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            for line in value.splitlines() or [""]:
                length = max(length, len(line))
        worksheet.column_dimensions[column_letter].width = min(max(length + 2, 12), 40)


def apply_wrap_text(worksheet, columns: list[int]) -> None:
    for column_index in columns:
        for row in worksheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")


def apply_reference_validations(workbook: Workbook, worksheet, row_limit: int) -> None:
    reference_sheet_name = f"'{REFERENCE_SHEET}'"
    validations = {
        "D": f"{reference_sheet_name}!$C$2:$C$3",
        "E": f"{reference_sheet_name}!$B$2:$B$3",
        "F": f"{reference_sheet_name}!$B$2:$B$3",
        "I": f"{reference_sheet_name}!$C$2:$C$3",
        "J": f"{reference_sheet_name}!$A$2:$A$5",
        "K": f"{reference_sheet_name}!$A$2:$A$5",
        "L": f"{reference_sheet_name}!$A$2:$A$5",
        "M": f"{reference_sheet_name}!$A$2:$A$5",
        "N": f"{reference_sheet_name}!$A$2:$A$5",
        "O": f"{reference_sheet_name}!$A$2:$A$5",
        "P": f"{reference_sheet_name}!$A$2:$A$5",
        "Q": f"{reference_sheet_name}!$A$2:$A$5",
    }
    for column_letter, formula in validations.items():
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.prompt = "Select one of the allowed values."
        validation.error = "Use one of the values listed on the Reference sheet."
        worksheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{row_limit}")

    category_validation = DataValidation(
        type="list",
        formula1=f"{reference_sheet_name}!$E$2:$E$500",
        allow_blank=False,
    )
    worksheet.add_data_validation(category_validation)
    category_validation.add(f"A2:A{row_limit}")


def build_reference_sheet(workbook: Workbook, sdk_values: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet(REFERENCE_SHEET)
    create_header_row(
        worksheet,
        ["presence", "geometry", "boolean", "severity", "categoryId", "serviceId"],
    )
    max_rows = max(
        len(PRESENCE_VALUES),
        len(GEOMETRY_VALUES),
        len(BOOLEAN_VALUES),
        len(SEVERITY_VALUES),
        len(sdk_values["categoryIds"]),
        len(sdk_values["services"]),
    )
    for index in range(max_rows):
        worksheet.append(
            [
                PRESENCE_VALUES[index] if index < len(PRESENCE_VALUES) else None,
                GEOMETRY_VALUES[index] if index < len(GEOMETRY_VALUES) else None,
                BOOLEAN_VALUES[index] if index < len(BOOLEAN_VALUES) else None,
                SEVERITY_VALUES[index] if index < len(SEVERITY_VALUES) else None,
                sdk_values["categoryIds"][index] if index < len(sdk_values["categoryIds"]) else None,
                sdk_values["services"][index] if index < len(sdk_values["services"]) else None,
            ]
        )
    auto_fit_columns(worksheet)
    worksheet.sheet_state = "hidden"


def build_instructions_sheet(workbook: Workbook) -> None:
    worksheet = workbook.active
    worksheet.title = INSTRUCTIONS_SHEET
    worksheet["A1"] = "WME Place Harmonizer config workbook"
    worksheet["A1"].font = Font(bold=True, size=14)
    rows = [
        "Gebruik de andere tabbladen om de config te bewerken.",
        "Lege cellen verwijderen een veld uit de JSON, behalve wanneer 'isDefined' op true staat.",
        "Gebruik in lijstvelden zoals services en geometryAllowed één waarde per regel in dezelfde cel.",
        "Editor notes staan op een apart tabblad met één regel per note en taal.",
        "Het tabblad 'Extra JSON' bewaart onderdelen die niet netjes in vaste kolommen passen.",
        "Import gebruikt de workbook als bron van waarheid en schrijft daarna weer nette JSON weg.",
    ]
    for index, line in enumerate(rows, start=3):
        worksheet[f"A{index}"] = line
    worksheet.column_dimensions["A"].width = 110


def export_overview_sheet(workbook: Workbook, config: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet(OVERVIEW_SHEET)
    create_header_row(worksheet, ["path", "label", "value", "encoding"])
    for path, label in OVERVIEW_FIELDS:
        value = get_nested(config, path)
        encoding, serialized = encode_value(value)
        worksheet.append([path, label, serialized, encoding])
    apply_wrap_text(worksheet, [3])
    auto_fit_columns(worksheet)


def export_rules_sheet(workbook: Workbook, config: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet(RULES_SHEET)
    create_header_row(worksheet, ["ruleId", "enabled", "severity", "extraJson"])
    for rule_id, rule in (config.get("rules") or {}).items():
        extra = {
            key: value
            for key, value in rule.items()
            if key not in {"enabled", "severity"}
        }
        worksheet.append(
            [
                rule_id,
                stringify_boolean(rule.get("enabled")),
                rule.get("severity"),
                json.dumps(extra, ensure_ascii=False) if extra else None,
            ]
        )
    auto_fit_columns(worksheet)


def export_formatting_sheet(workbook: Workbook, config: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet(FORMATTING_SHEET)
    create_header_row(worksheet, ["section", "field", "value", "encoding"])
    formatting = config.get("formatting") or {}
    for section, values in formatting.items():
        if not isinstance(values, dict):
            encoding, serialized = encode_value(values)
            worksheet.append([section, "", serialized, encoding])
            continue
        for field, value in values.items():
            encoding, serialized = encode_value(value)
            worksheet.append([section, field, serialized, encoding])
    apply_wrap_text(worksheet, [3])
    auto_fit_columns(worksheet)


def export_category_sheet(workbook: Workbook, config: dict[str, Any], sdk_values: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet(CATEGORY_SHEET)
    create_header_row(worksheet, CATEGORY_COLUMNS)
    parents = build_category_parents(sdk_values)
    standards = config.get("categoryStandards") or {}

    for category_id in sdk_values["categoryIds"]:
        standard = standards.get(category_id)
        is_defined = category_id in standards
        geometry = standard.get("geometry") if isinstance(standard, dict) else None
        services = standard.get("services") if isinstance(standard, dict) else None
        address = standard.get("address") if isinstance(standard, dict) else None
        extra = {}
        if isinstance(standard, dict):
            extra = {
                key: value
                for key, value in standard.items()
                if key
                not in {
                    "geometry",
                    "lockLevel",
                    "cityInVenueName",
                    "phone",
                    "url",
                    "openingHours",
                    "navigationPoints",
                    "externalProviderIds",
                    "services",
                    "address",
                    "editorNotes",
                }
            }

        worksheet.append(
            [
                category_id,
                parents.get(category_id),
                "main"
                if category_id in sdk_values["mainCategories"]
                else "subcategory"
                if parents.get(category_id)
                else "standalone",
                stringify_boolean(is_defined),
                geometry.get("required") if isinstance(geometry, dict) else None,
                geometry.get("recommended") if isinstance(geometry, dict) else None,
                join_lines(geometry.get("allowed")) if isinstance(geometry, dict) else None,
                standard.get("lockLevel") if isinstance(standard, dict) else None,
                stringify_boolean(standard.get("cityInVenueName")) if isinstance(standard, dict) else None,
                standard.get("phone") if isinstance(standard, dict) else None,
                standard.get("url") if isinstance(standard, dict) else None,
                standard.get("openingHours") if isinstance(standard, dict) else None,
                standard.get("navigationPoints") if isinstance(standard, dict) else None,
                standard.get("externalProviderIds") if isinstance(standard, dict) else None,
                address.get("city") if isinstance(address, dict) else None,
                address.get("street") if isinstance(address, dict) else None,
                address.get("houseNumber") if isinstance(address, dict) else None,
                join_lines(services.get("required")) if isinstance(services, dict) else None,
                join_lines(services.get("recommended")) if isinstance(services, dict) else None,
                join_lines(services.get("discouraged")) if isinstance(services, dict) else None,
                join_lines(services.get("forbidden")) if isinstance(services, dict) else None,
                json.dumps(extra, ensure_ascii=False) if extra else None,
            ]
        )

    apply_wrap_text(worksheet, [7, 18, 19, 20, 21, 22])
    apply_reference_validations(workbook, worksheet, len(sdk_values["categoryIds"]) + 20)
    auto_fit_columns(worksheet)


def export_editor_notes_sheet(workbook: Workbook, config: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet(EDITOR_NOTES_SHEET)
    create_header_row(worksheet, ["categoryId", "locale", "noteIndex", "note"])
    standards = config.get("categoryStandards") or {}
    for category_id, standard in standards.items():
        editor_notes = standard.get("editorNotes") if isinstance(standard, dict) else None
        if not isinstance(editor_notes, dict):
            continue
        for locale, notes in editor_notes.items():
            for note_index, note in enumerate(notes, start=1):
                worksheet.append([category_id, locale, note_index, note])
    apply_wrap_text(worksheet, [4])
    auto_fit_columns(worksheet)


def export_extra_json_sheet(workbook: Workbook, config: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet(EXTRA_JSON_SHEET)
    create_header_row(worksheet, ["path", "jsonValue"])
    remainder = remove_managed_paths(config)
    for key, value in remainder.items():
        worksheet.append([key, json.dumps(value, ensure_ascii=False, indent=2)])
    apply_wrap_text(worksheet, [2])
    auto_fit_columns(worksheet)


def export_workbook(config_path: Path, output_path: Path) -> None:
    root = config_path.parent.parent.parent if config_path.parent.name in {"countries", "communities", "states"} else config_path.parent.parent
    if not (root / "reference" / "sdk-values.json").exists():
        root = config_path.parent.parent

    config = load_json(config_path)
    sdk_values = load_sdk_values(root)

    workbook = Workbook()
    build_instructions_sheet(workbook)
    export_overview_sheet(workbook, config)
    export_rules_sheet(workbook, config)
    export_formatting_sheet(workbook, config)
    export_category_sheet(workbook, config, sdk_values)
    export_editor_notes_sheet(workbook, config)
    export_extra_json_sheet(workbook, config)
    build_reference_sheet(workbook, sdk_values)
    workbook.save(output_path)


def import_overview(workbook) -> dict[str, Any]:
    worksheet = workbook[OVERVIEW_SHEET]
    config: dict[str, Any] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        path, _label, value, encoding = row
        if not path:
            continue
        decoded = decode_value(encoding, value)
        if decoded is None:
            continue
        set_nested(config, str(path), decoded)
    return config


def import_rules(workbook) -> dict[str, Any]:
    worksheet = workbook[RULES_SHEET]
    rules: dict[str, Any] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        rule_id, enabled, severity, extra_json = row
        if not read_cell(rule_id):
            continue
        rule: dict[str, Any] = {}
        parsed_enabled = parse_boolean(enabled)
        if parsed_enabled is not None:
            rule["enabled"] = parsed_enabled
        parsed_severity = read_cell(severity)
        if parsed_severity is not None:
            rule["severity"] = parsed_severity
        extra = parse_json_cell(extra_json)
        if isinstance(extra, dict):
            rule.update(extra)
        if rule:
            rules[str(rule_id)] = rule
    return rules


def import_formatting(workbook) -> dict[str, Any]:
    worksheet = workbook[FORMATTING_SHEET]
    formatting: dict[str, Any] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        section, field, value, encoding = row
        section_name = read_cell(section)
        field_name = read_cell(field)
        if not section_name:
            continue
        decoded = decode_value(encoding, value)
        if decoded is None:
            continue
        if field_name:
            bucket = formatting.setdefault(section_name, {})
            bucket[field_name] = decoded
        else:
            formatting[section_name] = decoded
    return formatting


def import_category_standards(workbook) -> dict[str, Any]:
    worksheet = workbook[CATEGORY_SHEET]
    headers = [cell.value for cell in worksheet[1]]
    standards: dict[str, Any] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        category_id = read_cell(data.get("categoryId"))
        if not category_id:
            continue

        is_defined = parse_boolean(data.get("isDefined"))
        geometry: dict[str, Any] = {}
        geometry_required = read_cell(data.get("geometryRequired"))
        geometry_recommended = read_cell(data.get("geometryRecommended"))
        geometry_allowed = split_lines(data.get("geometryAllowed"))
        if geometry_required is not None:
            geometry["required"] = geometry_required
        if geometry_recommended is not None:
            geometry["recommended"] = geometry_recommended
        if geometry_allowed:
            geometry["allowed"] = geometry_allowed

        services: dict[str, Any] = {}
        for field_name, column_name in [
            ("required", "servicesRequired"),
            ("recommended", "servicesRecommended"),
            ("discouraged", "servicesDiscouraged"),
            ("forbidden", "servicesForbidden"),
        ]:
            values = split_lines(data.get(column_name))
            if values:
                services[field_name] = values

        address: dict[str, Any] = {}
        for field_name, column_name in [
            ("city", "addressCity"),
            ("street", "addressStreet"),
            ("houseNumber", "addressHouseNumber"),
        ]:
            value = read_cell(data.get(column_name))
            if value is not None:
                address[field_name] = value

        standard: dict[str, Any] = {}
        if geometry:
            standard["geometry"] = geometry

        lock_level = parse_integer(data.get("lockLevel"))
        if lock_level is not None:
            standard["lockLevel"] = lock_level

        for field_name in [
            "cityInVenueName",
            "phone",
            "url",
            "openingHours",
            "navigationPoints",
            "externalProviderIds",
        ]:
            raw_value = data.get(field_name)
            if field_name == "cityInVenueName":
                value = parse_boolean(raw_value)
            else:
                value = read_cell(raw_value)
            if value is not None:
                standard[field_name] = value

        if services:
            standard["services"] = services
        if address:
            standard["address"] = address

        extra = parse_json_cell(data.get("extraJson"))
        if isinstance(extra, dict):
            standard.update(extra)

        if is_defined is True or standard:
            standards[category_id] = standard
    return standards


def import_editor_notes(workbook, standards: dict[str, Any]) -> None:
    worksheet = workbook[EDITOR_NOTES_SHEET]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        category_id, locale, _note_index, note = row
        category_name = read_cell(category_id)
        locale_code = read_cell(locale)
        note_text = read_cell(note)
        if not category_name or not locale_code or not note_text:
            continue
        standard = standards.setdefault(category_name, {})
        editor_notes = standard.setdefault("editorNotes", {})
        notes = editor_notes.setdefault(locale_code, [])
        notes.append(note_text)


def import_extra_json(workbook, config: dict[str, Any]) -> None:
    worksheet = workbook[EXTRA_JSON_SHEET]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        path, json_value = row
        target_path = read_cell(path)
        if not target_path:
            continue
        parsed = parse_json_cell(json_value)
        if parsed is None:
            continue
        existing = get_nested(config, target_path)
        if isinstance(existing, dict) and isinstance(parsed, dict):
            set_nested(config, target_path, deep_merge(existing, parsed))
            continue
        set_nested(config, target_path, parsed)


def import_workbook(workbook_path: Path, output_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    config = import_overview(workbook)

    rules = import_rules(workbook)
    if rules:
        config["rules"] = rules

    formatting = import_formatting(workbook)
    if formatting:
        config["formatting"] = formatting

    standards = import_category_standards(workbook)
    import_editor_notes(workbook, standards)
    if standards:
        config["categoryStandards"] = standards

    import_extra_json(workbook, config)
    write_json(output_path, config)


def roundtrip_test(config_path: Path) -> None:
    original = load_json(config_path)
    with tempfile.TemporaryDirectory(prefix="config-excel-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        workbook_path = temp_dir / "config.xlsx"
        imported_path = temp_dir / "imported.json"
        export_workbook(config_path, workbook_path)
        import_workbook(workbook_path, imported_path)
        imported = load_json(imported_path)

    if imported != original:
        print("ERROR Excel round-trip produced a different config", file=sys.stderr)
        print("Original:", json.dumps(original, indent=2, ensure_ascii=False), file=sys.stderr)
        print("Imported:", json.dumps(imported, indent=2, ensure_ascii=False), file=sys.stderr)
        raise SystemExit(1)

    print(f"OK round-trip preserved {config_path}")


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Export/import WME Place Harmonizer config files to and from Excel."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    export_parser = subparsers.add_parser("export", help="Export a JSON config to .xlsx")
    export_parser.add_argument("config", type=Path, help="Path to the source JSON config")
    export_parser.add_argument("output", type=Path, help="Path to the target .xlsx file")

    import_parser = subparsers.add_parser("import", help="Import a .xlsx workbook into JSON")
    import_parser.add_argument("workbook", type=Path, help="Path to the source .xlsx file")
    import_parser.add_argument("output", type=Path, help="Path to the target JSON config")

    roundtrip_parser = subparsers.add_parser(
        "roundtrip", help="Verify export/import preserves a config exactly"
    )
    roundtrip_parser.add_argument("config", type=Path, help="Path to the JSON config to test")

    return parser


def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()

    if args.command == "export":
        export_workbook(args.config.resolve(), args.output.resolve())
        print(f"OK exported {args.config} to {args.output}")
        return

    if args.command == "import":
        import_workbook(args.workbook.resolve(), args.output.resolve())
        print(f"OK imported {args.workbook} to {args.output}")
        return

    if args.command == "roundtrip":
        roundtrip_test(args.config.resolve())
        return

    parser.error(f"Unsupported command {args.command}")


if __name__ == "__main__":
    main()
